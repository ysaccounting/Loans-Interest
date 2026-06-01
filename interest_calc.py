"""
Daily-balance interest calculator.

Reads one or more QuickBooks-style Transaction Reports (.xlsx) where each
sub-account block has a Beginning Balance row followed by dated transactions
with a running Balance column, and produces ONE workbook containing:

    * a "Loan Terms" tab documenting each rate block + notes, and
    * one calculation tab per uploaded report.

The loan each report belongs to is auto-detected from its content (company
name + account labels), so file names do not matter.

Interest method (daily balances):
    daily_rate   = annual_rate / days_in_year
    # days(row)  = next_txn_date - this_txn_date
                   (the final row of a block fills out the rest of the month
                    so the days in a block always sum to days_in_month)
    interest(row)= daily_rate * (sign * balance) * # days

`sign` is chosen automatically per tab so the total interest is always
positive, even when the running balances are negative.
"""

import calendar
import datetime as dt
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONTHS = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}

# ----- loan-type configuration ------------------------------------------
# group -> default annual rate
RATE_GROUP_DEFAULTS = {"affiliates": 0.085, "ystix": 0.12, "mazel": 0.13}

# group -> (label shown on Loan Terms, up-to-3 notes)
GROUP_TERMS = {
    "affiliates": ("Affiliates", ["", "", ""]),
    "ystix": ("YS Tix - YS Affiliates", [
        "USE FOR INTEREST CHARGED BY Y&S TICKETS TO YS AFFILIATES LLC",
        "Y&S Tickets - Debit due from/to YS Affiliates, Credit Interest Expense Offset (YS Affil)",
        "YS Affiliates - Debit Interest Income Offset (YS Affil), Credit due from/to Y&S Tickets",
    ]),
    "mazel": ("Mazel (Damona & Crew)", [
        "ONLY FOR DAMONA AND CREW. INTEREST INCOME GOES ON Mazel Investing",
        "", "",
    ]),
}
GROUP_ORDER = ["affiliates", "ystix", "mazel"]

# loan type -> {tab, group}
LOAN_TYPES = {
    "affiliates":  {"tab": "Affiliates",            "group": "affiliates"},
    "ystix":       {"tab": "YS Tix - YS Affiliates", "group": "ystix"},
    "damona":      {"tab": "Mazel - Damona",         "group": "mazel"},
    "eitz_chaim":  {"tab": "Mazel Eitz Chaim, TV",   "group": "mazel"},
}

# ----- styling (neutral; no yellow highlight, no blue font) --------------
FONT_NAME = "Arial"
HDR_FILL = PatternFill("solid", start_color="F2F2F2")
ACCT_FILL = PatternFill("solid", start_color="F2F2F2")
TOTAL_FILL = PatternFill("solid", start_color="E6E6E6")
GRAND_FILL = PatternFill("solid", start_color="D9D9D9")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '$#,##0.00;($#,##0.00);"-"'
DAYS_FMT = "0"
DATE_FMT = "mm/dd/yyyy"


# ----- parsing -----------------------------------------------------------
def _to_date(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        t = value.replace(",", "").replace("$", "").strip()
        if t in ("", "-"):
            return None
        try:
            return float(t)
        except ValueError:
            return None
    return None


def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 40) + 1):
        for c in range(1, min(ws.max_column, 15) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == "date":
                return r, c
    return 5, 2


def _detect_period(ws):
    for r in range(1, 8):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                m = re.search(r"([A-Za-z]+)\s+(\d{4})", v)
                if m and m.group(1).lower() in MONTHS:
                    return MONTHS[m.group(1).lower()], int(m.group(2))
    return None, None


def parse_report(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    company = ws.cell(row=1, column=1).value or "Company"
    title = ws.cell(row=2, column=1).value or "Transaction Report"
    period_txt = ws.cell(row=3, column=1).value or ""

    hdr_row, date_col = _find_header_row(ws)
    cols = {
        "date": date_col, "type": date_col + 1, "num": date_col + 2,
        "adj": date_col + 3, "name": date_col + 4, "memo": date_col + 5,
        "account": date_col + 6, "split": date_col + 7,
        "amount": date_col + 8, "balance": date_col + 9,
    }

    month, year = _detect_period(ws)

    blocks = []
    labels = []          # every column-A label, for loan detection
    current = None
    pending_label = None

    def close():
        nonlocal current
        if current and (current["rows"] or current["beginning_balance"] is not None):
            blocks.append(current)
        current = None

    for r in range(hdr_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        a = a.strip() if isinstance(a, str) else a
        b = ws.cell(row=r, column=cols["date"]).value
        b_str = b.strip() if isinstance(b, str) else b

        if a:
            low = a.lower()
            labels.append(a)
            if low.startswith("total for") or low == "total":
                close()
                pending_label = None
                continue
            close()
            pending_label = a
            continue

        if isinstance(b_str, str) and "beginning balance" in b_str.lower():
            if current is None:
                current = {"account": pending_label or "Account",
                           "rows": [], "beginning_balance": None}
            current["beginning_balance"] = _num(ws.cell(row=r, column=cols["balance"]).value) or 0.0
            continue

        d = _to_date(b_str)
        if d is not None:
            if current is None:
                current = {"account": pending_label or "Account",
                           "rows": [], "beginning_balance": None}
            current["rows"].append({
                "date": d,
                "type": ws.cell(row=r, column=cols["type"]).value,
                "num": ws.cell(row=r, column=cols["num"]).value,
                "adj": ws.cell(row=r, column=cols["adj"]).value,
                "name": ws.cell(row=r, column=cols["name"]).value,
                "memo": ws.cell(row=r, column=cols["memo"]).value,
                "account": ws.cell(row=r, column=cols["account"]).value,
                "split": ws.cell(row=r, column=cols["split"]).value,
                "amount": _num(ws.cell(row=r, column=cols["amount"]).value),
                "balance": _num(ws.cell(row=r, column=cols["balance"]).value),
            })
    close()

    if month and year:
        days_in_month = calendar.monthrange(year, month)[1]
    else:
        all_dates = [row["date"] for blk in blocks for row in blk["rows"]]
        if all_dates:
            ref = max(all_dates)
            month, year = ref.month, ref.year
            days_in_month = calendar.monthrange(year, month)[1]
        else:
            days_in_month = 30

    return {"company": company, "title": title, "period": period_txt,
            "month": month, "year": year, "days_in_month": days_in_month,
            "blocks": blocks, "labels": labels}


# ----- loan-type detection (from data, not file name) --------------------
def detect_loan_type(parsed):
    parts = [str(parsed.get("company") or "")]
    parts += [str(x) for x in parsed.get("labels", [])]
    for blk in parsed["blocks"]:
        parts.append(str(blk.get("account") or ""))
        for rec in blk["rows"]:
            parts += [str(rec.get("account") or ""), str(rec.get("name") or ""),
                      str(rec.get("memo") or "")]
    blob = " ".join(parts).lower()

    if "eitz chaim" in blob or "ticket vault" in blob or "subnotes" in blob:
        return "eitz_chaim"
    if "damona" in blob:
        return "damona"
    if ("y&s tickets" in blob or "ys tickets" in blob or "y&s tix" in blob
            or "ys tix" in blob or "tickets" in blob):
        return "ystix"
    if "affiliates" in blob or "notes receivable" in blob:
        return "affiliates"
    return "affiliates"


# ----- numeric computation (also used to pick the sign) ------------------
def compute_interest(parsed, annual_rate, days_in_year, sign=1):
    daily = annual_rate / days_in_year
    dim = parsed["days_in_month"]
    month = parsed["month"] or 4
    year = parsed["year"] or dt.date.today().year
    first = dt.date(year, month, 1)

    breakdown, grand = [], 0.0
    for blk in parsed["blocks"]:
        beg = blk["beginning_balance"]
        if beg is None and blk["rows"]:
            f = blk["rows"][0]
            beg = (f["balance"] - f["amount"]) if (f["balance"] is not None
                   and f["amount"] is not None) else 0.0
        seq = []
        if beg is not None:
            seq.append((first, beg))
        for rec in blk["rows"]:
            seq.append((rec["date"],
                        rec["balance"] if rec["balance"] is not None else beg))
        if not seq:
            continue

        total, cum = 0.0, 0
        for i, (d, bal) in enumerate(seq):
            days = (seq[i + 1][0] - d).days if i < len(seq) - 1 else dim - cum
            cum += days
            total += daily * (sign * (bal or 0.0)) * days
        breakdown.append({"account": blk["account"], "interest": round(total, 2)})
        grand += total
    return breakdown, round(grand, 2)


def pick_sign(parsed, annual_rate, days_in_year):
    """Choose the sign that makes the total interest non-negative."""
    _, total = compute_interest(parsed, annual_rate, days_in_year, 1)
    return -1 if total < 0 else 1


# ----- Loan Terms sheet (hardcoded "Summary of Loans and Terms") ---------
# Each block writes: (lender_label, lender), (borrower_label, borrower),
# "Terms", rate, # days, daily rate, then one or more "QBO Notes" lines.
# Block N starts at row 3 + 9*N; the calc tab for that loan links its
# interest formula to the block's daily-rate cell ($B$<daily_row>).
LOAN_TERM_BLOCKS = [
    {
        "loan_type": "affiliates", "rate_group": "affiliates",
        "lender_label": "Lender", "lender": "YS Affiliates LLC",
        "borrower_label": "Borrowers", "borrower": "Multiple Affiliates",
        "notes": ["create invoices on YS Affiliates LLC and and bills on the Affiliates"],
    },
    {
        "loan_type": "ystix", "rate_group": "ystix",
        "lender_label": "Lender", "lender": "Y&S Tickets Inc",
        "borrower_label": "Borrower", "borrower": "YS Affiliates LLC",
        "notes": [
            "Y&S Tickets - Debit due from/to YS Affiliates, Credit Interest Expense Offset (YS Affil)",
            "YS Affiliates - Debit Interest Income Offset (YS Affil), Credit due from/to Y&S Tickets",
        ],
    },
    {
        "loan_type": "damona", "rate_group": "mazel",
        "lender_label": "Lender", "lender": "Mazel Investing",
        "borrower_label": "Borrower", "borrower": "Damona & Crew",
        "notes": ["create invoice on Mazel Investing and bill on Damona & Crew"],
    },
    {
        "loan_type": "eitz_chaim", "rate_group": "mazel",
        "lender_label": "Lenders", "lender": "Eitz Chaim, TicketVault",
        "borrower_label": "Borrower", "borrower": "Mazel Investing",
        "notes": ["create bills on Mazel Investing"],
    },
]


def _build_loan_terms(wb, company, rates, days_in_year):
    """
    Write the hardcoded "Summary of Loans and Terms" tab.
    Returns {loan_type: "$B$<daily_rate_row>"} for the calc-tab formulas.
    """
    ws = wb.active
    ws.title = "Loan Terms"
    ws["A1"] = "Summary of Loans and Terms"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=13)

    daily_cells = {}
    r = 3
    for blk in LOAN_TERM_BLOCKS:
        rate = rates.get(blk["rate_group"], RATE_GROUP_DEFAULTS[blk["rate_group"]])

        ws.cell(row=r, column=1, value=blk["lender_label"])
        ws.cell(row=r, column=2, value=blk["lender"])
        ws.cell(row=r + 1, column=1, value=blk["borrower_label"])
        ws.cell(row=r + 1, column=2, value=blk["borrower"])
        ws.cell(row=r + 2, column=1, value="Terms")

        rate_row, days_row, daily_row = r + 3, r + 4, r + 5
        ws.cell(row=rate_row, column=1, value="Annual Interest Rate").font = \
            Font(name=FONT_NAME, italic=True)
        ws.cell(row=rate_row, column=2, value=rate).number_format = "0.000%"
        ws.cell(row=days_row, column=1, value="# Days in Year")
        ws.cell(row=days_row, column=2, value=days_in_year).number_format = "0"
        ws.cell(row=daily_row, column=1, value="Daily Interest Rate")
        ws.cell(row=daily_row, column=2,
                value=f"=B{rate_row}/B{days_row}").number_format = "0.00000000%"

        note_row = r + 6
        ws.cell(row=note_row, column=1, value="QBO Notes")
        for i, note in enumerate(blk["notes"]):
            ws.cell(row=note_row + i, column=2, value=note)

        daily_cells[blk["loan_type"]] = f"$B${daily_row}"
        # Next block starts after the last note line + 2 blank rows, matching
        # the target spacing (blocks at rows 3, 12, 22, 31).
        r = note_row + len(blk["notes"]) + 2

    ws.column_dimensions["A"].width = 23.4
    ws.column_dimensions["B"].width = 17.4
    ws.column_dimensions["C"].width = 36.4
    ws.column_dimensions["D"].width = 72.0
    return daily_cells


# ----- Calculation sheet -------------------------------------------------
def _build_calc_sheet(wb, parsed, tab_name, daily_cell, sign):
    month = parsed["month"] or 4
    year = parsed["year"] or dt.date.today().year
    first_of_month = dt.datetime(year, month, 1)

    ws = wb.create_sheet(tab_name[:31])
    ws["A1"] = parsed["company"]
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=12)
    ws["A2"] = parsed["title"]
    ws["A3"] = parsed["period"]
    ws["M1"] = "# Days in Month"
    ws["N1"] = parsed["days_in_month"]
    ws["M1"].font = Font(name=FONT_NAME, bold=True)
    ws["N1"].font = Font(name=FONT_NAME)

    headers = ["Date", "Transaction Type", "Num", "Adj", "Name",
               "Memo/Description", "Account", "Split", "Amount", "Balance",
               "# Days", "Interest", "# Days (Cum)", "Cash Transferred?"]
    hdr_row = 5
    for i, h in enumerate(headers):
        c = ws.cell(row=hdr_row, column=2 + i, value=h)
        c.font = Font(name=FONT_NAME, bold=True)
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER

    COL = {name: get_column_letter(2 + i) for i, name in enumerate(
        ["date", "type", "num", "adj", "name", "memo", "account", "split",
         "amount", "balance", "days", "interest", "cum", "cash"])}
    sign_prefix = "-" if sign < 0 else ""

    def write_row(r, date_val, rec, bal_val):
        ws.cell(row=r, column=2, value=date_val).number_format = DATE_FMT
        if rec:
            ws.cell(row=r, column=3, value=rec.get("type"))
            ws.cell(row=r, column=4, value=rec.get("num"))
            ws.cell(row=r, column=5, value=rec.get("adj"))
            ws.cell(row=r, column=6, value=rec.get("name"))
            ws.cell(row=r, column=7, value=rec.get("memo"))
            ws.cell(row=r, column=8, value=rec.get("account"))
            ws.cell(row=r, column=9, value=rec.get("split"))
            ws.cell(row=r, column=10, value=rec.get("amount")).number_format = MONEY
        ws.cell(row=r, column=11, value=bal_val).number_format = MONEY
        days = ws.cell(row=r, column=12)
        days.value = (f'=IF({COL["balance"]}{r+1}<>"",'
                      f'{COL["date"]}{r+1}-{COL["date"]}{r},'
                      f'$N$1-{COL["cum"]}{r-1})')
        days.number_format = DAYS_FMT
        intr = ws.cell(row=r, column=13)
        intr.value = (f"='Loan Terms'!{daily_cell}*{sign_prefix}"
                      f"{COL['balance']}{r}*{COL['days']}{r}")
        intr.number_format = MONEY
        intr.font = Font(name=FONT_NAME)
        ws.cell(row=r, column=14,
                value=f"={COL['days']}{r}+{COL['cum']}{r-1}").number_format = DAYS_FMT
        for col in range(2, 16):
            ws.cell(row=r, column=col).border = BORDER

    sum_cells = []
    row = hdr_row + 2
    for blk in parsed["blocks"]:
        lbl = ws.cell(row=row, column=1, value=blk["account"])
        lbl.font = Font(name=FONT_NAME, bold=True)
        lbl.fill = ACCT_FILL
        row += 1
        block_start = row

        beg = blk["beginning_balance"]
        if beg is None and blk["rows"]:
            f = blk["rows"][0]
            beg = (f["balance"] - f["amount"]) if (f["balance"] is not None
                   and f["amount"] is not None) else 0.0
        seq = []
        if beg is not None:
            seq.append((first_of_month, None, beg))
        for rec in blk["rows"]:
            seq.append((dt.datetime.combine(rec["date"], dt.time()), rec,
                        rec["balance"] if rec["balance"] is not None else beg))
        if not seq:
            continue
        for date_val, rec, bal_val in seq:
            write_row(row, date_val, rec, bal_val)
            row += 1

        tot = row
        ws.cell(row=tot, column=7, value=f"Total — {blk['account']}").font = \
            Font(name=FONT_NAME, bold=True)
        ws.cell(row=tot, column=12,
                value=f"=SUM({COL['days']}{block_start}:{COL['days']}{tot-1})"
                ).number_format = DAYS_FMT
        ws.cell(row=tot, column=13,
                value=f"=SUM({COL['interest']}{block_start}:{COL['interest']}{tot-1})"
                ).number_format = MONEY
        for cc in (12, 13):
            ws.cell(row=tot, column=cc).font = Font(name=FONT_NAME, bold=True)
            ws.cell(row=tot, column=cc).fill = TOTAL_FILL
        ws.cell(row=tot, column=7).fill = TOTAL_FILL
        sum_cells.append(f"{COL['interest']}{tot}")
        row += 2

    g = row + 1
    ws.cell(row=g, column=7, value="TOTAL MONTHLY INTEREST").font = \
        Font(name=FONT_NAME, bold=True, size=11)
    gt = ws.cell(row=g, column=13,
                 value=("=" + "+".join(sum_cells)) if sum_cells else 0)
    gt.number_format = MONEY
    gt.font = Font(name=FONT_NAME, bold=True, size=11)
    gt.fill = GRAND_FILL
    ws.cell(row=g, column=7).fill = GRAND_FILL

    widths = {"A": 22, "B": 11, "C": 15, "D": 6, "E": 6, "F": 16, "G": 34,
              "H": 24, "I": 14, "J": 14, "K": 15, "L": 8, "M": 14, "N": 12,
              "O": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B6"


# ----- orchestrator ------------------------------------------------------
def generate(jobs, output_path, rates=None, days_in_year=365):
    """
    jobs: list of {"path": str, "loan_type": <key>|"auto"|""}.
    rates: {"affiliates":0.085, "ystix":0.12, "mazel":0.13} (by group, decimals).
    Builds ONE workbook (Loan Terms + a tab per job) and returns a summary dict.
    """
    rates = rates or {}
    company = "Company"
    year = dt.date.today().year

    prepared = []
    for job in jobs:
        parsed = parse_report(job["path"])
        lt = job.get("loan_type")
        if lt not in LOAN_TYPES:
            lt = detect_loan_type(parsed)
        group = LOAN_TYPES[lt]["group"]
        rate = rates.get(group, RATE_GROUP_DEFAULTS[group])
        sign = pick_sign(parsed, rate, days_in_year)
        prepared.append({"parsed": parsed, "loan_type": lt, "group": group,
                         "rate": rate, "sign": sign})

    if prepared:
        company = prepared[0]["parsed"]["company"]

    wb = Workbook()
    group_rates = {g: rates.get(g, RATE_GROUP_DEFAULTS[g]) for g in GROUP_ORDER}
    daily_cells = _build_loan_terms(wb, company, group_rates, days_in_year)

    used_names, tabs = set(), []
    for p in prepared:
        cfg = LOAN_TYPES[p["loan_type"]]
        tab = cfg["tab"]
        n = 2
        while tab[:31] in used_names:
            tab = f"{cfg['tab']} {n}"
            n += 1
        used_names.add(tab[:31])

        _build_calc_sheet(wb, p["parsed"], tab, daily_cells[p["loan_type"]], p["sign"])
        breakdown, grand = compute_interest(
            p["parsed"], p["rate"], days_in_year, p["sign"])
        tabs.append({
            "tab_name": tab, "loan_type": p["loan_type"],
            "company": p["parsed"]["company"], "period": p["parsed"]["period"],
            "days_in_month": p["parsed"]["days_in_month"],
            "rate_pct": round(p["rate"] * 100, 4),
            "num_accounts": len(p["parsed"]["blocks"]),
            "breakdown": breakdown, "grand_total": grand,
        })

    wb.save(output_path)
    return {"company": company, "tabs": tabs}


if __name__ == "__main__":
    import sys
    print(generate([{"path": sys.argv[1], "loan_type": "auto"}], sys.argv[2]))
